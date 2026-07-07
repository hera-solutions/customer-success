#!/usr/bin/env python3
"""
Convert an Amazon DSP vehicle export (xlsx) into the Hera Vehicles import CSV.

Mapping rules are documented in vehicles-amazon-dsp-mapping.md. Update both
files together if Amazon or Hera changes their schema.

Usage:
    python3 vehicles-amazon-dsp-convert.py \\
        --input  /path/to/VehiclesData.xlsx \\
        --output /path/to/vehicles-hera-import.csv
"""

import argparse
import csv
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


HERA_HEADERS = [
    "Status ( Active / Inactive / Inactive - Maintenance / Inactive - Grounded )",
    " Vehicle Name",
    " License Plate",
    " License Plate Expiration",
    " State",
    " VIN",
    " Gas Card ( Last 6 )",
    " Mileage",
    " Parking Space",
    " Start Date",
    " End Date",
    " Ownership ( Lease / Own / Rent )",
    " Vehicle Type",
    "Description",
    " Make",
    " Model",
    " Year",
    " Company ( Alamo / Avis / Budget / Dollar / Element / Enterprise / Fluid Truck / Hertz / Lease Plan / Merchants Fleet / National / Thrifty / Zeeba )",
    " Rent Agreement Number",
]

SERVICE_TYPE_MAP = {
    "Standard Parcel - Extra Large Van - US": "Standard Parcel XL",
    "Standard Parcel - Large Van": "Standard Parcel Large",
    "Standard Parcel - Medium Van": "Standard Parcel",
    "Standard Parcel - Small Van": "Standard Parcel Small",
    "Standard Parcel - Custom Delivery Van 14ft": "Custom Delivery Van",
    "Standard Parcel - Custom Delivery Van 16ft": "Custom Delivery Van",
    "Standard Parcel Electric - Rivian MEDIUM": "EDV",
    "Box Truck Parcel (Large)": "10,000lbs Van",
}

SERVICE_TIER_FALLBACK = {
    "EXTRA_LARGE_CARGO_VAN": "Standard Parcel XL",
    "LARGE_CARGO_VAN": "Standard Parcel Large",
    "STANDARD_CARGO_VAN": "Standard Parcel",
    "SMALL_CARGO_VAN": "Standard Parcel Small",
    "CUSTOM_DELIVERY_VAN_FOURTEEN_FT": "Custom Delivery Van",
    "CUSTOM_DELIVERY_VAN_SIXTEEN_FT": "Custom Delivery Van",
    "ELECTRIC_RPV_MEDIUM": "EDV",
    "BOX_TRUCK_LARGE": "10,000lbs Van",
}

PROVIDER_MAP = {
    "BUDGET": "Budget",
    "ELEMENT": "Element",
    "Enterprise": "Enterprise",
    "LP": "Lease Plan",
    "U Haul": "U-Haul",
    "MERCHANTS": "Merchants Fleet",
    "Zeeba": "Zeeba",
    "Kingbee": "Kingbee",
    "Ryder": "Ryder",
}

OWNERSHIP_MAP = {
    "RENTAL": "Rent",
    "AMAZON_OWNED": "Lease",
    "AMAZON_RENTAL": "Lease",
    "LEASE": "Lease",
    "SELF_OWNED": "Own",
}


def convert_date(s):
    if not s:
        return ""
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return s


def map_status(status, op_status):
    if op_status == "GROUNDED":
        return "Inactive - Grounded"
    if status == "ACTIVE" and op_status in ("OPERATIONAL", "READY_FOR_AUDIT"):
        return "Active"
    if status == "INACTIVE":
        return "Inactive"
    return ""


def map_vehicle_type(service_type, service_tier):
    if service_type and service_type in SERVICE_TYPE_MAP:
        return SERVICE_TYPE_MAP[service_type]
    if service_tier and service_tier in SERVICE_TIER_FALLBACK:
        return SERVICE_TIER_FALLBACK[service_tier]
    return ""


def extract_state(s):
    if not s:
        return ""
    return s.split(" - ")[0].strip()


def convert(input_path: Path, output_path: Path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers)}

    converted = []
    notes = []

    for row_num, r in enumerate(rows[1:], start=2):
        def g(h):
            return r[idx[h]] if h in idx else None

        vin = g("vin")
        plate = g("licensePlateNumber")
        name = g("vehicleName") or ""
        service_type = g("serviceType")
        service_tier = g("serviceTier")
        veh_type = map_vehicle_type(service_type, service_tier)

        if not service_type and service_tier:
            notes.append(
                f"Row {row_num} (VIN {vin}): serviceType blank, "
                f"used serviceTier={service_tier} -> Vehicle Type='{veh_type}'"
            )
        if not name:
            notes.append(f"Row {row_num} (VIN {vin}, Plate {plate}): Vehicle Name blank")

        converted.append([
            map_status(g("status"), g("operationalStatus")),
            name,
            plate or "",
            convert_date(g("registrationExpiryDate")),
            extract_state(g("registeredState")),
            vin or "",
            "",  # Gas Card (Last 6)
            "",  # Mileage
            "",  # Parking Space
            convert_date(g("ownershipStartDate")),
            convert_date(g("ownershipEndDate")),
            OWNERSHIP_MAP.get(g("ownershipType"), ""),
            veh_type,
            "",  # Description
            g("make") or "",
            g("model") or "",
            g("year") or "",
            PROVIDER_MAP.get(g("vehicleProvider"), g("vehicleProvider") or ""),
            "",  # Rent Agreement Number
        ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(HERA_HEADERS)
        w.writerows(converted)

    return converted, notes


def validate(converted):
    issues = []
    vins = [r[5] for r in converted]
    plates = [r[2] for r in converted]
    vin_dupes = [v for v, c in Counter(vins).items() if c > 1]
    plate_dupes = [p for p, c in Counter(plates).items() if c > 1]
    if vin_dupes:
        issues.append(f"DUPLICATE VINs (block import): {vin_dupes}")
    if plate_dupes:
        issues.append(f"DUPLICATE License Plates (block import): {plate_dupes}")
    bad_vins = [(i + 2, v) for i, v in enumerate(vins) if len(str(v)) != 17]
    if bad_vins:
        issues.append(f"VINs not 17 chars: {bad_vins}")
    required_cols = {0: "Status", 2: "License Plate", 4: "State", 5: "VIN",
                     9: "Start Date", 11: "Ownership", 12: "Vehicle Type",
                     14: "Make", 15: "Model", 16: "Year"}
    for col_idx, col_name in required_cols.items():
        blanks = [i + 2 for i, r in enumerate(converted) if not r[col_idx]]
        if blanks:
            issues.append(f"Blank {col_name} in rows: {blanks}")
    return issues


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    args = p.parse_args()

    converted, notes = convert(args.input, args.output)
    print(f"Wrote {len(converted)} rows to {args.output}")

    if notes:
        print("\n=== Per-row notes ===")
        for n in notes:
            print(n)

    issues = validate(converted)
    print("\n=== Validation ===")
    if issues:
        for i in issues:
            print(f"  [!] {i}")
        sys.exit(1)
    else:
        print("  All checks passed.")


if __name__ == "__main__":
    main()
